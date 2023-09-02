
import getpass
import winreg
import openpyxl
import hashlib
import pandas as pd
from openpyxl.styles import PatternFill

def calculate_hash(value, hash_algorithm):
    hasher = hashlib.new(hash_algorithm)
    hasher.update(value.encode('utf-8'))
    return hasher.hexdigest()

def get_startup_entries():
    startup_entries = []

    try:
        admin_reg_path = r"SOFTWARE\Microsoft\Windows\CurrentVersion\Run"
        with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, admin_reg_path, 0, winreg.KEY_READ | winreg.KEY_WOW64_64KEY) as admin_reg_key:
            for i in range(winreg.QueryInfoKey(admin_reg_key)[1]):
                name, value, _ = winreg.EnumValue(admin_reg_key, i)
                hash_value = calculate_hash(value, 'sha256')
                startup_entries.append((getpass.getuser(), 'Administrator', name, value, hash_value))
    except Exception as e:
        print("Error retrieving administrator startup entries:", str(e))

    try:
        user_reg_path = r"SOFTWARE\Microsoft\Windows\CurrentVersion\Run"
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, user_reg_path, 0, winreg.KEY_READ) as user_reg_key:
            for i in range(winreg.QueryInfoKey(user_reg_key)[1]):
                name, value, _ = winreg.EnumValue(user_reg_key, i)
                hash_value = calculate_hash(value, 'sha256')
                startup_entries.append((getpass.getuser(), 'User', name, value, hash_value))
    except Exception as e:
        print("Error retrieving user startup entries:", str(e))

    return startup_entries

def compare_hash_with_input(input_file, startup_entries):
    matched_entries = []

    with open(input_file, 'r') as file:
        given_hashes = [line.strip() for line in file.readlines()]

    for given_hash in given_hashes:
        for entry in startup_entries:
            if entry[4] == given_hash:
                matched_entries.append(entry)

    return matched_entries

startup_entries = get_startup_entries()

input_file = 'input.txt'

matched_entries = compare_hash_with_input(input_file, startup_entries)

# Create a DataFrame for all startup entries
df_all_entries = pd.DataFrame(startup_entries, columns=['User Name', 'User Type', 'Name', 'Path', 'Hash Value'])

# Add a new column 'Matched' and initialize it with False
df_all_entries['Matched'] = False

# Iterate over the DataFrame and update 'Matched' column for highlighted entries
for index, row in df_all_entries.iterrows():
    hash_value = row['Hash Value']
    if any(entry[4] == hash_value for entry in matched_entries):
        df_all_entries.at[index, 'Matched'] = True

# Save all entries DataFrame to Excel
df_all_entries.to_excel('all_entries.xlsx', index=False)

# Load the workbook
workbook = openpyxl.load_workbook('all_entries.xlsx')

# Select the active sheet
sheet = workbook.active

# Define the fill color for matched entries (customize here)
fill_color = "FFFF00"  # Specify the desired color code (e.g., "FFFF00" for yellow)

# Create a PatternFill object with the specified color
matched_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

# Iterate over the rows and highlight the matched entries with the custom fill color
for row in sheet.iter_rows(min_row=2):
    hash_value = row[4].value
    if any(entry[4] == hash_value for entry in matched_entries):
        for cell in row:
            cell.fill = matched_fill

# Save the workbook with the highlighted matched entries
workbook.save('all_entries.xlsx')

print('All entries with user names, hash values, and matching status are stored in all_entries.xlsx.')
